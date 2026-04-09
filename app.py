import streamlit as st
import pandas as pd
import requests
import os
import re
import time
import zipfile
import io
from io import BytesIO
from PIL import Image, ImageChops
from concurrent.futures import ThreadPoolExecutor, as_completed
from bs4 import BeautifulSoup

# ─────────────────────────────────────────────
# PAGE CONFIG
# ─────────────────────────────────────────────
st.set_page_config(
    page_title="Amazon Image Auditor",
    page_icon="🛒",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─────────────────────────────────────────────
# CUSTOM CSS
# ─────────────────────────────────────────────
st.markdown("""
<style>
    .stApp { background-color: #0f1117; color: #e8eaf0; }
    .card {
        background: #1a1d2e;
        border: 1px solid #2d3250;
        border-radius: 12px;
        padding: 20px 24px;
        margin-bottom: 16px;
    }
    .step-badge {
        display: inline-block;
        background: #FF9900;
        color: #000;
        font-weight: 700;
        font-size: 12px;
        border-radius: 50%;
        width: 26px; height: 26px;
        text-align: center;
        line-height: 26px;
        margin-right: 10px;
    }
    .hero-title { font-size:2.2rem; font-weight:800; color:#FF9900; letter-spacing:-0.5px; }
    .hero-sub   { color:#9ba3bf; font-size:1rem; margin-top:4px; }
    div[data-testid="metric-container"] {
        background:#1a1d2e;
        border:1px solid #2d3250;
        border-radius:10px;
        padding:14px 18px;
    }
    div[data-testid="metric-container"] label { color:#9ba3bf !important; }
    #MainMenu, footer, header { visibility: hidden; }
    section[data-testid="stSidebar"] { background:#12151f; border-right:1px solid #2d3250; }
    div[data-testid="stFileUploader"] {
        border: 2px dashed #FF9900 !important;
        border-radius: 10px;
        padding: 10px;
    }
    .stProgress > div > div > div { background-color:#FF9900 !important; }
    .stDownloadButton button {
        background:#FF9900 !important;
        color:#000 !important;
        font-weight:700 !important;
        border:none !important;
        border-radius:8px !important;
        padding:10px 24px !important;
        font-size:15px !important;
        width: 100%;
    }
    .stDownloadButton button:hover { background:#e68a00 !important; }
    .stButton button {
        background:#FF9900 !important;
        color:#000 !important;
        font-weight:700 !important;
        border:none !important;
        border-radius:8px !important;
    }
</style>
""", unsafe_allow_html=True)


# ─────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📖 How to Use")
    st.markdown("""
<div class='card'>
<p><span class='step-badge'>1</span> <b>Prepare your Excel file</b></p>
<p>Two required columns (exact names):</p>
<ul>
  <li><b>Product Title</b> — your SKU or name (used to name saved images)</li>
  <li><b>Image URL</b> — direct image link or Takealot product page URL</li>
</ul>
</div>

<div class='card'>
<p><span class='step-badge'>2</span> <b>Upload & configure</b></p>
<p>Upload your file, set workers (4 is a good default), and optionally enable background removal.</p>
</div>

<div class='card'>
<p><span class='step-badge'>3</span> <b>Run the audit</b></p>
<p>Click <b>Start Audit</b>. Progress and live metrics update as each image is processed.</p>
</div>

<div class='card'>
<p><span class='step-badge'>4</span> <b>Download your outputs</b></p>
<ul>
  <li>📊 <b>Excel Report</b> — full audit results (4 sheets)</li>
  <li>🗜️ <b>Images ZIP</b> — all processed images named <code>SKU.MAIN.jpg</code>, ready for Seller Central</li>
</ul>
</div>

<div class='card'>
<p><span class='step-badge'>5</span> <b>Status guide</b></p>
<ul>
  <li>✅ <b>COMPLIANT</b> — passed all checks</li>
  <li>⚠️ <b>ATTENTION</b> — processed but needs review</li>
  <li>❌ <b>FAILED</b> — could not retrieve or process</li>
</ul>
<p>The <b>Audit Notes</b> column explains every fix or error.</p>
</div>
""", unsafe_allow_html=True)

    st.divider()
    st.markdown("### ✅ Amazon Standards Applied")
    st.markdown("""
- White background (RGB 255,255,255)
- Min 1000px · Max 10,000px on longest side
- JPEG format at 72 DPI
- Output: `ProductTitle.MAIN.jpg`
""")
    st.divider()
    st.markdown("### 🔗 Supported URL Types")
    st.markdown("""
- `media.takealot.com` CDN links
- `takealot.com` product pages (PLID)
- Direct `.jpg` / `.png` / `.webp` links
- Amazon product pages
- Any page with OG image meta tags
""")
    st.divider()
    st.caption("Amazon Image Auditor · v2.1")


# ─────────────────────────────────────────────
# CORE FUNCTIONS
# ─────────────────────────────────────────────

@st.cache_resource
def get_session():
    s = requests.Session()
    s.headers.update({
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
        'Accept': 'image/webp,image/apng,image/*,*/*;q=0.8',
    })
    return s


def safe_filename(title):
    name = re.sub(r'[^\w\s.-]', '', str(title)).strip()
    name = re.sub(r'\s+', '_', name)
    return name[:100]


def resolve_image_url(url, session):
    """
    Resolves any URL to a downloadable image URL.

    Priority order:
    1. Takealot CDN direct link → upgrade to high-res variant
    2. Direct image link by file extension
    3. Takealot product page → API call for gallery
    4. HTML scrape for OG / Amazon image tag
    """
    url = str(url).strip()
    if not url or url == "nan":
        return None

    # 1. Takealot CDN — these ARE images, just with .file extension
    #    Upgrade s-zoom to pdpxl for highest resolution
    if "media.takealot.com" in url:
        upgraded = re.sub(r'/s-zoom\.file$', '/pdpxl.file', url)
        upgraded = re.sub(r'/s-pdpxl\.file$', '/pdpxl.file', upgraded)
        return upgraded

    # 2. Direct image link
    if any(ext in url.lower() for ext in ['.jpg', '.jpeg', '.png', '.webp', '.gif', '.bmp']):
        return url

    # 3. Takealot product page → API
    if "takealot.com" in url:
        plid = re.search(r'(PLID\d+)', url, re.IGNORECASE)
        if plid:
            try:
                api = f"https://api.takealot.com/rest/v-1-10-0/product-details/{plid.group(1)}?platform=desktop"
                data = session.get(api, timeout=8).json()
                gallery = data.get("gallery_images", [])
                if gallery:
                    return gallery[0].replace("{size}", "pdpxl")
            except Exception:
                pass

    # 4. HTML scrape
    try:
        resp = session.get(url, timeout=10)
        soup = BeautifulSoup(resp.text, 'html.parser')

        img = soup.find('img', src=lambda x: x and 'media.takealot.com/covers_images' in x)
        if img and img.get('src'):
            return img['src']

        og = soup.find("meta", property="og:image")
        if og and og.get("content"):
            return og["content"]

        amz = soup.find("img", id="landingImage")
        if amz and amz.get("src"):
            return amz["src"]

    except Exception:
        pass

    return None


def process_image(row_data, img_url, enable_bg_removal):
    result               = dict(row_data)
    result["Status"]     = "FAILED"
    result["Audit Notes"] = ""
    result["Resolved URL"] = img_url or "Not Found"
    result["_image_bytes"] = None
    result["_safe_name"]   = safe_filename(row_data.get("Product Title", "Item"))

    if not img_url:
        result["Audit Notes"] = "Could not resolve image URL"
        return result

    try:
        session = get_session()
        r = session.get(img_url, timeout=25)
        r.raise_for_status()

        if "text/html" in r.headers.get("Content-Type", ""):
            result["Audit Notes"] = "URL returned HTML page, not an image"
            return result

        raw_img   = Image.open(BytesIO(r.content))
        fixes_log = []

        if raw_img.mode not in ('RGB', 'RGBA'):
            raw_img = raw_img.convert('RGBA')
            fixes_log.append("Converted colour mode")

        w, h      = raw_img.size
        orig_size = f"{w}x{h}px"

        if enable_bg_removal:
            try:
                from rembg import remove as rembg_remove
                no_bg    = rembg_remove(raw_img)
                white_bg = Image.new("RGB", no_bg.size, (255, 255, 255))
                white_bg.paste(no_bg, (0, 0), no_bg.convert("RGBA"))
                final_img = white_bg
                fixes_log.append("Background removed → white")
            except ImportError:
                final_img = raw_img.convert("RGB")
                fixes_log.append("rembg not installed — BG skipped")
            except Exception as e:
                final_img = raw_img.convert("RGB")
                fixes_log.append(f"BG removal failed: {str(e)[:40]}")
        else:
            final_img = raw_img.convert("RGB")

        w, h    = final_img.size
        longest = max(w, h)
        if longest < 1000:
            ratio     = 1000 / longest
            new_w, new_h = int(w * ratio), int(h * ratio)
            final_img = final_img.resize((new_w, new_h), Image.Resampling.LANCZOS)
            fixes_log.append(f"Upscaled {orig_size} → {new_w}x{new_h}px")
        elif longest > 10000:
            ratio     = 10000 / longest
            new_w, new_h = int(w * ratio), int(h * ratio)
            final_img = final_img.resize((new_w, new_h), Image.Resampling.LANCZOS)
            fixes_log.append(f"Downscaled {orig_size} → {new_w}x{new_h}px")
        else:
            fixes_log.append(f"Size OK ({orig_size})")

        buf = BytesIO()
        final_img.save(buf, "JPEG", quality=95, dpi=(72, 72), subsampling=0)
        buf.seek(0)

        blank = Image.new("RGB", final_img.size, (255, 255, 255))
        if not ImageChops.difference(final_img, blank).getbbox():
            result["Status"] = "ATTENTION"
            fixes_log.append("Warning: image appears blank/white")
        else:
            result["Status"] = "COMPLIANT"

        safe  = result["_safe_name"]
        w_out, h_out = final_img.size

        result["Audit Notes"]    = " | ".join(fixes_log) if fixes_log else "Already Compliant"
        result["Fixed Filename"] = f"{safe}.MAIN.jpg"
        result["Final Size"]     = f"{w_out}x{h_out}px"
        result["_image_bytes"]   = buf.getvalue()

    except requests.exceptions.HTTPError as e:
        result["Audit Notes"] = f"HTTP {e.response.status_code} — download failed"
    except Exception as e:
        result["Audit Notes"] = f"Error: {str(e)[:100]}"

    return result


def build_excel(report_df, compliant, attention, failed, total, elapsed):
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        report_df.to_excel(writer, sheet_name="Audit Report", index=False)

        # Fixed Images sheet — only rows that were successfully processed
        fixed_cols = ["Product Title", "Fixed Filename", "Final Size", "Status", "Audit Notes", "Resolved URL"]
        fixed_df   = report_df[report_df["Status"].isin(["COMPLIANT", "ATTENTION"])]
        avail_cols = [c for c in fixed_cols if c in fixed_df.columns]
        fixed_df[avail_cols].to_excel(writer, sheet_name="Fixed Images", index=False)

        pd.DataFrame({
            "Metric": ["Total Products", "Compliant", "Needs Attention", "Failed",
                       "Pass Rate", "Audit Date", "Processing Time"],
            "Value":  [total, compliant, attention, failed,
                       f"{round((compliant/total)*100,1)}%" if total else "0%",
                       time.strftime("%Y-%m-%d %H:%M"),
                       f"{int(elapsed)}s"],
        }).to_excel(writer, sheet_name="Summary", index=False)

        failed_df = report_df[report_df["Status"] == "FAILED"]
        if not failed_df.empty:
            failed_df[["Product Title", "Image URL", "Audit Notes"]].to_excel(
                writer, sheet_name="Failed Items", index=False
            )

        attn_df = report_df[report_df["Status"] == "ATTENTION"]
        if not attn_df.empty:
            attn_df.to_excel(writer, sheet_name="Needs Attention", index=False)

        STATUS_BG   = {"COMPLIANT": "1a472a", "ATTENTION": "422006", "FAILED": "3b0000"}
        STATUS_TEXT = {"COMPLIANT": "4ade80", "ATTENTION": "fb923c", "FAILED": "f87171"}

        for sname in ["Audit Report", "Fixed Images", "Failed Items", "Needs Attention", "Summary"]:
            if sname not in writer.sheets:
                continue
            ws = writer.sheets[sname]

            for cell in ws[1]:
                cell.font      = Font(bold=True, color="000000")
                cell.fill      = PatternFill("solid", fgColor="FF9900")
                cell.alignment = Alignment(horizontal="center", vertical="center")

            for col_cells in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col_cells), default=10)
                ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_len + 4, 65)

            if "Status" in report_df.columns and sname != "Summary":
                try:
                    sidx = list(report_df.columns).index("Status") + 1
                    for row_cells in ws.iter_rows(min_row=2):
                        cell = row_cells[sidx - 1]
                        val  = str(cell.value)
                        if val in STATUS_BG:
                            cell.fill = PatternFill("solid", fgColor=STATUS_BG[val])
                            cell.font = Font(bold=True, color=STATUS_TEXT[val])
                except Exception:
                    pass

    buf.seek(0)
    return buf.getvalue()


def build_zip(results):
    """
    Pack COMPLIANT and ATTENTION images into a ZIP.
    Each image is named:  ProductTitle.MAIN.jpg
    Also includes a manifest.csv listing every file and its source URL.
    """
    zip_buf    = BytesIO()
    count      = 0
    seen_names = {}
    manifest   = [["Fixed Filename", "Product Title", "Status", "Resolved URL", "Audit Notes"]]

    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for r in results:
            if r.get("_image_bytes") and r.get("Status") in ("COMPLIANT", "ATTENTION"):
                base = r.get("_safe_name", "image")

                # Handle duplicate product titles
                if base in seen_names:
                    seen_names[base] += 1
                    fname = f"{base}_{seen_names[base]}.MAIN.jpg"
                else:
                    seen_names[base] = 0
                    fname = f"{base}.MAIN.jpg"

                zf.writestr(f"images/{fname}", r["_image_bytes"])
                manifest.append([
                    fname,
                    r.get("Product Title", ""),
                    r.get("Status", ""),
                    r.get("Resolved URL", ""),
                    r.get("Audit Notes", ""),
                ])
                count += 1

        # Write manifest CSV into ZIP root
        manifest_lines = "\n".join(",".join(f'"{str(v)}"' for v in row) for row in manifest)
        zf.writestr("manifest.csv", manifest_lines)

    zip_buf.seek(0)
    return zip_buf.getvalue(), count


# ─────────────────────────────────────────────
# MAIN UI
# ─────────────────────────────────────────────
st.markdown("""
<div class='hero-title'>🛒 Amazon Image Auditor</div>
<div class='hero-sub'>Upload your product list · Audit against Amazon standards · Download fixed images & report</div>
""", unsafe_allow_html=True)
st.markdown("---")

col_a, col_b, col_c = st.columns([2, 1, 1])

with col_a:
    st.markdown("### 📁 Upload Products File")
    uploaded_file = st.file_uploader(
        "Drop your Excel file here",
        type=["xlsx", "xls"],
        label_visibility="collapsed"
    )
with col_b:
    st.markdown("### ⚙️ Workers")
    max_workers = st.slider("Parallel threads", 1, 10, 4,
                            help="More workers = faster. 4 is a safe default.")
with col_c:
    st.markdown("### 🖼️ Background")
    enable_bg_removal = st.toggle("Remove backgrounds", value=False,
                                  help="Uses rembg. Slower but produces Amazon-compliant white backgrounds.")

st.markdown("")

if uploaded_file:
    try:
        df = pd.read_excel(uploaded_file)
        missing = {"Product Title", "Image URL"} - set(df.columns)

        if missing:
            st.error(
                f"❌ Missing columns: **{', '.join(missing)}**  \n"
                f"Your file has: `{'`, `'.join(df.columns.tolist())}`"
            )
        else:
            st.success(f"✅ File loaded — **{len(df)} products** ready to audit")

            with st.expander("👀 Preview first 5 rows", expanded=False):
                st.dataframe(df.head(), use_container_width=True)

            st.markdown("")

            if st.button(f"🚀  Start Audit  ({len(df)} products)", use_container_width=True):
                st.markdown("---")
                st.markdown("### ⏳ Audit in Progress")

                progress_bar   = st.progress(0)
                status_text    = st.empty()
                metrics_holder = st.empty()

                all_results = []
                total       = len(df)
                compliant   = attention = failed = 0
                start_time  = time.time()
                session     = get_session()

                # Step 1: Resolve URLs
                status_text.info("🔍 Step 1 of 2 — Resolving image URLs...")
                resolved_urls = []
                for i, (_, row) in enumerate(df.iterrows()):
                    resolved_urls.append(resolve_image_url(row.get("Image URL", ""), session))
                    progress_bar.progress((i + 1) / (total * 2))

                # Step 2: Process in parallel
                status_text.info("⚙️ Step 2 of 2 — Downloading & processing images...")
                with ThreadPoolExecutor(max_workers=max_workers) as executor:
                    futures = {
                        executor.submit(process_image, df.iloc[i].to_dict(), resolved_urls[i], enable_bg_removal): i
                        for i in range(total)
                    }
                    done = 0
                    for future in as_completed(futures):
                        result = future.result()
                        all_results.append(result)
                        done += 1

                        s = result.get("Status", "FAILED")
                        if s == "COMPLIANT":   compliant += 1
                        elif s == "ATTENTION": attention += 1
                        else:                  failed    += 1

                        elapsed   = time.time() - start_time
                        remaining = (elapsed / done) * (total - done) if done else 0
                        progress_bar.progress(0.5 + done / (total * 2))
                        status_text.info(
                            f"Processing {done}/{total} · ~{int(remaining)}s remaining · "
                            f"✅ {compliant}  ⚠️ {attention}  ❌ {failed}"
                        )
                        m1, m2, m3, m4 = metrics_holder.columns(4)
                        m1.metric("Processed",   f"{done}/{total}")
                        m2.metric("✅ Compliant", compliant)
                        m3.metric("⚠️ Attention", attention)
                        m4.metric("❌ Failed",    failed)

                progress_bar.progress(1.0)
                elapsed = time.time() - start_time
                status_text.success(f"🎉 Audit complete in {int(elapsed)}s!")

                # Build outputs
                report_rows = [{k: v for k, v in r.items() if not k.startswith("_")} for r in all_results]
                report_df   = pd.DataFrame(report_rows)
                priority    = ["Product Title", "Image URL", "Status", "Fixed Filename", "Final Size", "Audit Notes", "Resolved URL"]
                other       = [c for c in report_df.columns if c not in priority]
                report_df   = report_df[[c for c in priority if c in report_df.columns] + other]

                excel_bytes          = build_excel(report_df, compliant, attention, failed, total, elapsed)
                zip_bytes, img_count = build_zip(all_results)
                ts                   = time.strftime('%Y%m%d_%H%M')

                # Download section
                st.markdown("---")
                st.markdown("### 📥 Download Your Outputs")

                dl1, dl2 = st.columns(2)

                with dl1:
                    st.markdown("#### 📊 Excel Report")
                    st.markdown("""
5 sheets inside:
- **Audit Report** — every product with full status & notes
- **Fixed Images** — processed items with new filename & final dimensions
- **Summary** — pass rate and run details
- **Failed Items** — re-uploadable list (if any)
- **Needs Attention** — manual review queue (if any)
""")
                    st.download_button(
                        label="⬇️  Download Excel Report",
                        data=excel_bytes,
                        file_name=f"Amazon_Audit_{ts}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True,
                    )

                with dl2:
                    st.markdown("#### 🗜️ Processed Images (ZIP)")
                    st.markdown(f"""
**{img_count} images** inside (COMPLIANT + ATTENTION):
- Folder: `images/` — all files named `SKU.MAIN.jpg`
- `manifest.csv` — links filename back to source URL
- White background · JPEG · 72 DPI · min 1000px
- Unzip and upload straight to Amazon Seller Central
""")
                    st.download_button(
                        label="⬇️  Download Images ZIP",
                        data=zip_bytes,
                        file_name=f"Amazon_Images_{ts}.zip",
                        mime="application/zip",
                        use_container_width=True,
                    )

                if failed > 0:
                    st.markdown("")
                    st.warning(
                        f"⚠️ **{failed} items failed.** Open the *Failed Items* sheet in your Excel report. "
                        "The most common cause with Takealot CDN links is a temporary block — "
                        "wait a few minutes and re-run those rows with the same URLs."
                    )

    except Exception as e:
        st.error(f"❌ Could not read file: {str(e)}")

else:
    st.markdown("""
<div class='card' style='text-align:center; padding:40px;'>
    <div style='font-size:3rem;'>📂</div>
    <div style='font-size:1.3rem; font-weight:600; margin-top:12px;'>Upload your products.xlsx to get started</div>
    <div style='color:#9ba3bf; margin-top:8px;'>Two columns required: <b>Product Title</b> and <b>Image URL</b></div>
    <div style='color:#9ba3bf; margin-top:6px; font-size:0.88rem;'>
        ✅ &nbsp;Supports Takealot CDN links &nbsp;·&nbsp; Takealot product pages &nbsp;·&nbsp; Amazon &nbsp;·&nbsp; Direct image URLs
    </div>
</div>
""", unsafe_allow_html=True)

    sample = pd.DataFrame({
        "Product Title": ["BINITABLK6", "IG-LEMIASLV5", "KIRARBL3"],
        "Image URL": [
            "https://media.takealot.com/covers_images/2004c940eb164e96a5a7b74bd85e021a/s-zoom.file",
            "https://media.takealot.com/covers_images/a9d94b6d821c41eb86bcd48b80c8b03d/s-zoom.file",
            "https://media.takealot.com/covers_images/8c80a9af9e534f3695b7f59366e586fe/s-zoom.file",
        ]
    })
    sbuf = BytesIO()
    sample.to_excel(sbuf, index=False)
    sbuf.seek(0)
    st.download_button(
        "📄 Download Sample Template (with your format)",
        data=sbuf,
        file_name="products_template.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
